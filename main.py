from flask import Flask, request, jsonify
import openpyxl
import random

app = Flask(__name__)

# Загружаем Excel-файл
workbook = openpyxl.load_workbook("questions.xlsx")

# Получаем имена листов (названия — это кнопки)
sheet_names = workbook.sheetnames

# Загружаем вопросы по листам
quizzes = {}
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        question, options, correct, explanation = row
        data.append({
            "Вопрос": str(question).strip(),
            "Варианты": [opt.strip() for opt in str(options).split(";")],
            "Правильный": str(correct).strip(),
            "Пояснение": str(explanation).strip()
        })
    quizzes[sheet_name] = data


def get_random_question(sheet_name):
    """Возвращает случайный вопрос из выбранного листа"""
    return random.choice(quizzes[sheet_name])


@app.route("/", methods=["POST"])
def main():
    req = request.json
    session = req.get("session", {})
    state = req.get("state", {}).get("session", {})
    command = req["request"]["command"].strip().lower()

    response = {
        "version": req["version"],
        "session": req["session"],
        "response": {"end_session": False}
    }

    # === 1. Приветствие ===
    if session.get("new", False):
        response["response"]["text"] = "Привет! Выбери документ для викторины 📘"
        response["response"]["buttons"] = [{"title": name} for name in sheet_names]
        return jsonify(response)

    # === 2. Пользователь выбрал лист ===
    if command.capitalize() in sheet_names:
        topic = command.capitalize()
        question = get_random_question(topic)
        response["response"]["text"] = (
            f'Вы выбрали "{topic}".\n\n'
            f'{question["Вопрос"]}\n'
            f'Варианты: {", ".join(question["Варианты"])}'
        )
        response["session_state"] = {"topic": topic, "question": question}
        return jsonify(response)

    # === 3. Пользователь отвечает на вопрос ===
    if state.get("question"):
        topic = state["topic"]
        question = state["question"]
        correct = question["Правильный"].lower()

        if command == correct.lower():
            next_q = get_random_question(topic)
            response["response"]["text"] = (
                f"Правильно! 🎉\n\n"
                f"Следующий вопрос:\n{next_q['Вопрос']}\n"
                f"Варианты: {', '.join(next_q['Варианты'])}"
            )
            response["session_state"] = {"topic": topic, "question": next_q}
        else:
            response["response"]["text"] = (
                f"Неправильно 😢\n"
                f"{question['Пояснение']}\n\n"
                f"Хочешь попробовать другой документ?"
            )
            response["response"]["buttons"] = [{"title": name} for name in sheet_names]
            response["session_state"] = {}
        return jsonify(response)

    # === 4. Если непонятная команда ===
    response["response"]["text"] = "Я не поняла. Выбери документ:"
    response["response"]["buttons"] = [{"title": name} for name in sheet_names]
    return jsonify(response)


@app.route("/", methods=["GET"])
def ping():
    return "Навык Алисы работает!", 200


if __name__ == "__main__":
    print("🚀 Навык Алисы запущен на http://127.0.0.1:5000/")
    app.run(host="0.0.0.0", port=5000, debug=True)
