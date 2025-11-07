from flask import Flask, request, jsonify
import openpyxl
import random
import re
import os
import logging
from datetime import datetime

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 📂 Путь к Excel-файлу
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(BASE_DIR, "questions.xlsx")

if not os.path.exists(excel_path):
    raise FileNotFoundError(f"Файл {excel_path} не найден!")

# Загружаем Excel
workbook = openpyxl.load_workbook(excel_path)
sheet_names = workbook.sheetnames

# 🖼️ Словарь с ID картинок из каталога Алисы
ALICE_IMAGE_IDS = {
    "1": "997614/f3e84f7cd524f792e0c3",  # ВАША КАРТИНКА - вставлена сюда!
    # Добавьте другие картинки:
    # "2": "ваш_id_2",
    # "3": "ваш_id_3", ...
}


# ===============================
# 🔹 Подготовка базы вопросов
# ===============================
def parse_options(options_str):
    if not options_str:
        return []
    return [opt.strip() for opt in str(options_str).split(';') if opt.strip()]


def parse_correct(correct_str):
    if not correct_str:
        return []
    matches = re.findall(r'([А-ЯЁA-Z]\))', str(correct_str))
    return matches


def get_alice_image_id(image_name):
    """Получить ID картинки из каталога Алисы"""
    if not image_name:
        return None
    return ALICE_IMAGE_IDS.get(str(image_name).strip())


quizzes = {}
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        question, options, correct, explanation, image = (row + (None, None, None, None, None))[:5]
        if not question:
            continue

        alice_image_id = get_alice_image_id(image)

        data.append({
            "Вопрос": str(question).strip(),
            "Варианты": parse_options(options),
            "Правильный": parse_correct(correct),
            "Пояснение": str(explanation).strip() if explanation else "",
            "Изображение": alice_image_id
        })
    quizzes[sheet_name] = data


def get_random_question(topic, previous_questions=None):
    if topic not in quizzes or not quizzes[topic]:
        return None

    if previous_questions is None:
        previous_questions = []

    available_questions = [q for q in quizzes[topic] if q["Вопрос"] not in previous_questions]

    if not available_questions:
        available_questions = quizzes[topic]

    return random.choice(available_questions)


def normalize_answer(user_answer):
    if not user_answer:
        return ""

    user_answer = user_answer.strip().lower()

    digit_to_letter = {"1": "а", "2": "б", "3": "в", "4": "г", "5": "д", "6": "е"}
    if user_answer in digit_to_letter:
        return digit_to_letter[user_answer]

    user_answer = re.sub(r'[).\s,]', '', user_answer)

    if user_answer and user_answer[0] in 'абвгде':
        return user_answer[0]

    return ""


def normalize_correct_answers(correct_answers):
    normalized = []
    for answer in correct_answers:
        clean_answer = re.sub(r'[)\s]', '', answer).lower()
        if clean_answer and clean_answer[0] in 'абвгде':
            normalized.append(clean_answer[0])
    return normalized


def parse_multiple_answers(command):
    cleaned = re.sub(r'[.,;]', ' ', command.lower())
    answers = cleaned.split()

    normalized_answers = []
    valid_answers = set()

    for answer in answers:
        normalized = normalize_answer(answer)
        if normalized and normalized not in valid_answers:
            normalized_answers.append(normalized)
            valid_answers.add(normalized)

    return normalized_answers


# Временное хранилище сессий
user_sessions = {}


# ===============================
# 🚀 Основной Webhook
# ===============================
@app.route("/", methods=["POST"])
def main():
    try:
        req = request.json
        if not req:
            return jsonify_error("Пустой запрос")

        command = req["request"]["command"].strip().lower()
        session = req.get("session", {})
        session_id = session.get("session_id")

        logger.info(f"Получен запрос: команда='{command}', session_id={session_id}")

        user_state = user_sessions.get(session_id, {})

        response = {
            "version": req["version"],
            "session": req["session"],
            "response": {"end_session": False, "text": "", "buttons": []},
            "session_state": {}
        }

        # Новая сессия — приветствие
        if session.get("new", False):
            user_sessions[session_id] = {}
            buttons = [{"title": name} for name in sheet_names]
            response["response"]["text"] = "Привет! Выберите тему для тестирования:"
            response["response"]["buttons"] = buttons
            logger.info("Новая сессия: отправлено приветствие")
            return jsonify(response)

        # Назад в меню
        if any(nav_cmd in command for nav_cmd in ["назад", "меню", "главная", "выход"]):
            user_sessions[session_id] = {}
            buttons = [{"title": name} for name in sheet_names]
            response["response"]["text"] = "Вы вернулись в главное меню. Выберите тему:"
            response["response"]["buttons"] = buttons
            logger.info("Возврат в меню")
            return jsonify(response)

        # Пропуск вопроса
        if any(skip_cmd in command for skip_cmd in ["пропустить", "следующий", "дальше", "skip", "next"]):
            if user_state.get("mode") == "question" and user_state.get("topic"):
                topic = user_state["topic"]
                previous_questions = user_state.get("previous_questions", [])

                next_question = get_random_question(topic, previous_questions)
                if next_question:
                    options_text = "\n".join([f"{opt}" for opt in next_question["Варианты"]]) if next_question[
                        "Варианты"] else ""
                    response_text = (
                        f"Вопрос пропущен.\n\n"
                        f'Тема: "{topic}"\n\n'
                        f'{next_question["Вопрос"]}\n\n'
                        f'{options_text}'
                    )

                    if len(response_text) > 1000:
                        response_text = response_text[:997] + "..."

                    response["response"]["text"] = response_text

                    # Добавляем картинку если есть
                    if next_question["Изображение"]:
                        response["response"]["card"] = {
                            "type": "BigImage",
                            "image_id": next_question["Изображение"],
                            "title": "Изображение к вопросу",
                            "description": f"Тема: {topic}"
                        }

                    updated_previous_questions = previous_questions + [next_question["Вопрос"]]
                    user_sessions[session_id] = {
                        "topic": topic,
                        "question": next_question,
                        "previous_questions": updated_previous_questions,
                        "mode": "question"
                    }
                else:
                    response["response"]["text"] = "Вопросы в этой теме закончились."
                    user_sessions[session_id] = {}

                response["response"]["buttons"] = [
                    {"title": "Пропустить"},
                    {"title": "Назад в меню"}
                ]
                logger.info("Вопрос пропущен")
                return jsonify(response)

        # Помощь
        if command in ["помощь", "help", "что делать", "правила"]:
            if user_state.get("mode") == "question":
                response["response"]["text"] = (
                    f"Вы в режиме вопроса по теме '{user_state['topic']}'. "
                    f"Произнесите номер ответа (1-6) или букву (А-Е). "
                    f"Можно несколько ответов через пробел: '1 2' или 'а б'. "
                    f"Скажите 'пропустить' для перехода к следующему вопросу. "
                    f"Или скажите 'назад' для возврата в меню."
                )
            else:
                response["response"]["text"] = (
                    "Я помогу вам подготовиться к экзамену. "
                    "Выберите тему для тестирования или скажите 'назад' в любой момент. "
                    "Во время тестирования можно пропускать вопросы командой 'пропустить'."
                )
            response["response"]["buttons"] = [{"title": "Назад в меню"}]
            logger.info("Показана помощь")
            return jsonify(response)

        # Проверка выбора темы
        for sheet_name in sheet_names:
            if command == sheet_name.lower():
                topic = sheet_name
                question = get_random_question(topic)
                if not question:
                    response["response"]["text"] = f"В теме '{topic}' нет вопросов."
                    response["response"]["buttons"] = [{"title": "Назад в меню"}]
                    logger.warning(f"В теме '{topic}' нет вопросов")
                    return jsonify(response)

                options_text = "\n".join([f"{opt}" for opt in question["Варианты"]]) if question["Варианты"] else ""
                response_text = (
                    f'Тема: "{topic}"\n\n'
                    f'{question["Вопрос"]}\n\n'
                    f'{options_text}'
                )

                if len(response_text) > 1000:
                    response_text = response_text[:997] + "..."

                response["response"]["text"] = response_text

                # Добавляем картинку если есть
                if question["Изображение"]:
                    response["response"]["card"] = {
                        "type": "BigImage",
                        "image_id": question["Изображение"],
                        "title": "Изображение к вопросу",
                        "description": f"Тема: {topic}"
                    }

                response["response"]["buttons"] = [
                    {"title": "Пропустить"},
                    {"title": "Назад в меню"}
                ]

                user_sessions[session_id] = {
                    "topic": topic,
                    "question": question,
                    "previous_questions": [question["Вопрос"]],
                    "mode": "question"
                }

                logger.info(f"Выбрана тема '{topic}', сохранено состояние")
                return jsonify(response)

        # Ответ на вопрос
        if user_state.get("mode") == "question" and user_state.get("topic") and user_state.get("question"):
            topic = user_state["topic"]
            current_question = user_state["question"]
            previous_questions = user_state.get("previous_questions", [])

            logger.info(f"Обрабатываем ответ для темы '{topic}': '{command}'")

            user_answers = parse_multiple_answers(command)
            correct_answers_normalized = normalize_correct_answers(current_question["Правильный"])

            logger.info(f"Распознанные ответы: {user_answers}")
            logger.info(f"Правильные ответы: {correct_answers_normalized}")

            if not user_answers:
                response["response"]["text"] = (
                    f"Не понял ответ '{command}'. "
                    f"Используйте цифры 1-6 или буквы А-Е. "
                    f"Пример: '1', 'а', '1 2', 'а б'. "
                    f"Скажите 'пропустить' для перехода к следующему вопросу. "
                    f"Или скажите 'назад' для возврата в меню."
                )
                response["response"]["buttons"] = [
                    {"title": "Пропустить"},
                    {"title": "Назад в меню"}
                ]
                user_sessions[session_id] = user_state
                return jsonify(response)

            correct_given = [ans for ans in user_answers if ans in correct_answers_normalized]
            incorrect_given = [ans for ans in user_answers if ans not in correct_answers_normalized]

            if not incorrect_given and len(correct_given) == len(correct_answers_normalized):
                text = "Верно!"
            elif not incorrect_given and len(correct_given) > 0:
                missing = [ans for ans in correct_answers_normalized if ans not in user_answers]
                missing_text = ", ".join([f"{ans.upper()})" for ans in missing])
                text = f"Частично верно! Вы выбрали правильные ответы, но не хватает: {missing_text}\n\n{current_question['Пояснение']}"
            elif len(correct_given) > 0 and len(incorrect_given) > 0:
                correct_text = ", ".join([f"{ans.upper()})" for ans in correct_given])
                incorrect_text = ", ".join([f"{ans.upper()})" for ans in incorrect_given])
                text = f"Частично верно! Правильные: {correct_text}, неправильные: {incorrect_text}\n\n{current_question['Пояснение']}"
            else:
                correct_text = ", ".join(current_question["Правильный"])
                text = f"Неверно.\nПравильный ответ: {correct_text}\n\n{current_question['Пояснение']}"

            next_question = get_random_question(topic, previous_questions)
            if next_question:
                options_text = "\n".join([f"{opt}" for opt in next_question["Варианты"]]) if next_question[
                    "Варианты"] else ""
                text += (
                    f"\n\nСледующий вопрос:\n{next_question['Вопрос']}\n\n"
                    f"{options_text}"
                )
                if len(text) > 1000:
                    text = text[:997] + "..."

                # Добавляем картинку для следующего вопроса если есть
                if next_question["Изображение"]:
                    response["response"]["card"] = {
                        "type": "BigImage",
                        "image_id": next_question["Изображение"],
                        "title": "Изображение к вопросу",
                        "description": f"Тема: {topic}"
                    }

                updated_previous_questions = previous_questions + [next_question["Вопрос"]]
                user_sessions[session_id] = {
                    "topic": topic,
                    "question": next_question,
                    "previous_questions": updated_previous_questions,
                    "mode": "question"
                }
            else:
                text += "\n\nВопросы в этой теме закончились."
                user_sessions[session_id] = {}

            response["response"]["text"] = text
            response["response"]["buttons"] = [
                {"title": "Пропустить"},
                {"title": "Назад в меню"}
            ]
            return jsonify(response)

        # Если команда не распознана
        buttons = [{"title": name} for name in sheet_names]
        response["response"]["text"] = "Пожалуйста, выберите тему из предложенных ниже."
        response["response"]["buttons"] = buttons
        return jsonify(response)

    except Exception as e:
        logger.error(f"Ошибка обработки запроса: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify_error("Произошла ошибка. Пожалуйста, попробуйте еще раз.")


def jsonify_error(message):
    return jsonify({
        "version": "1.0",
        "response": {"text": message, "end_session": False},
        "session_state": {}
    })


@app.route("/", methods=["GET"])
def home():
    return jsonify({
        "status": "success",
        "message": "Навык Алисы работает.",
        "active_sessions": len(user_sessions),
        "topics_loaded": list(quizzes.keys())
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    logger.info(f"Запуск сервера на порту {port}")
    app.run(host="0.0.0.0", port=port, debug=False)